import os
from openpyxl import load_workbook, Workbook

class ExcelSchemaWriter:
    @classmethod
    def write_dynamic_data(cls, sheet_name, schema_fields, payload):
        """
        Parses lab_data.xlsx, searches for the indicated 'sheet_name'.
        Locates the exact header row by scanning for 'Sr.No' or 'Sr. No'.
        Gets the last Sr.No to auto-increment.
        Appends a new row ordering the payload data strictly based on the found Excel column indices.
        """
        excel_path = os.path.join(os.getcwd(), "lab_data.xlsx")
        
        # 1. Open or Create Workbook
        if not os.path.exists(excel_path):
            wb = Workbook()
            # Rename default sheet if empty
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
                 ws = wb.active
                 ws.title = sheet_name
            else:
                 ws = wb.create_sheet(title=sheet_name)
                 
            # If it's a completely fresh sheet, build the header row at Row 3 (matching template styling loosely)
            cls._create_fresh_header(ws, schema_fields)
            wb.save(excel_path)
            
        wb = load_workbook(excel_path)
        
        # 2. Try fetching the requested sheet, create if it doesn't exist
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            cls._create_fresh_header(ws, schema_fields)

        # 3. Locate the actual Header Row by searching for "Sr.No"
        header_row_idx = None
        sr_no_col_idx = None
        column_map = {} # { Excel Column Index (1-based) : Payload Key }

        for r_idx in range(1, 15): # Scan top 15 rows for the header
            for c_idx in range(1, 30): # Scan up to 30 columns
                cell_val = ws.cell(row=r_idx, column=c_idx).value
                if isinstance(cell_val, str):
                    val_clean = cell_val.strip().lower().replace(" ", "")
                    # "sr.no", "srno", "sr. no."
                    if val_clean in ["sr.no", "sr.no.", "srno", "srno."]:
                        header_row_idx = r_idx
                        sr_no_col_idx = c_idx
                        break
            if header_row_idx:
                break
                
        # If we still can't find a Sr.No, initialize a clean header at row 1
        if not header_row_idx:
            ws.delete_rows(1, ws.max_row)
            cls._create_fresh_header(ws, schema_fields)
            header_row_idx = 1
            sr_no_col_idx = 1

        # 4. Map the identified headers to Schema Keys
        # A cell might have trailing spaces, casing differences, etc.
        # So we do a normalized string comparison.
        def normalize_str(s):
            if not s: return ""
            return str(s).strip().lower().replace(" ", "").replace("_", "").replace(".", "")

        for f in schema_fields:
            lbl_norm = normalize_str(f["label"])
            # Scan the header row
            found_col = None
            max_col = ws.max_column or 30
            for c_idx in range(1, max_col + 1):
                 hd_val = ws.cell(row=header_row_idx, column=c_idx).value
                 if hd_val and normalize_str(hd_val) == lbl_norm:
                     found_col = c_idx
                     break
            
            if found_col:
                column_map[found_col] = f["key"]
            else:
                # If template is broken, auto-append the missing column header to the end
                new_col_idx = (ws.max_column or 0) + 1
                ws.cell(row=header_row_idx, column=new_col_idx, value=f["label"])
                column_map[new_col_idx] = f["key"]

        # 5. Determine the Auto-Incrementing Sr.No and next empty row
        last_sr = 0
        insert_row_idx = header_row_idx + 1

        # Scan downwards from header to find the last valid populated row
        for r_idx in range(header_row_idx + 1, ws.max_row + 10):
            val = ws.cell(row=r_idx, column=sr_no_col_idx).value
            if val is not None and str(val).strip() != "":
                try:
                    last_sr = int(val)
                except ValueError:
                    pass
                insert_row_idx = r_idx + 1
            elif val is None or str(val).strip() == "":
                 # Found empty row sequence, insertion point secured
                 break
                 
        next_sr = last_sr + 1

        # 6. Write Data
        # Set the Sr.No
        ws.cell(row=insert_row_idx, column=sr_no_col_idx, value=next_sr)
        
        # Set explicitly mapped columns
        for col_idx, expected_key in column_map.items():
            if expected_key in payload:
                ws.cell(row=insert_row_idx, column=col_idx, value=payload[expected_key])

        wb.save(excel_path)
        return True, next_sr

    @classmethod
    def _create_fresh_header(cls, ws, schema_fields):
        """Creates a default formatted header row starting at row 1 if template is missing."""
        ws.cell(row=1, column=1, value="Sr.No")
        for i, field in enumerate(schema_fields, start=2):
            ws.cell(row=1, column=i, value=field["label"])
