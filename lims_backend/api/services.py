from django.core.mail import send_mail
from django.conf import settings
import os
from openpyxl import load_workbook
import datetime

# ==========================================
# 1. EMAIL SERVICE
# ==========================================

class EmailService:
    # Use a real FROM email if you configure SMTP properly. 
    # For now, it will use DEFAULT_FROM_EMAIL from settings.
    FROM_EMAIL = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gsfc-lims.com')
    
    @classmethod
    def send_inventory_request_approved(cls, employee_email, material, quantity):
        subject = "Inventory Request Approved"
        message = f"Hello Employee,\n\nYour request for material \"{material} - {quantity}\" has been approved by the admin.\n\nRegards\nGSFC LIMS System"
        try:
            send_mail(subject, message, cls.FROM_EMAIL, [employee_email], fail_silently=True)
        except Exception as e:
            print(f"Failed to send approval email: {e}")

    @classmethod
    def send_inventory_request_rejected(cls, employee_email, material, quantity):
        subject = "Inventory Request Rejected"
        message = f"Hello Employee,\n\nYour request for material \"{material} - {quantity}\" has been rejected by the admin.\n\nRegards\nGSFC LIMS System"
        try:
            send_mail(subject, message, cls.FROM_EMAIL, [employee_email], fail_silently=True)
        except Exception as e:
            print(f"Failed to send rejection email: {e}")

    @classmethod
    def send_low_stock_alert(cls, admin_email, material, current_balance, threshold):
        subject = "ALERT: Low Inventory Stock"
        message = f"ALERT Admin,\n\nThe inventory for \"{material}\" has fallen to {current_balance}, which is below the minimum threshold of {threshold}.\nPlease reorder immediately.\n\nRegards\nGSFC LIMS System"
        try:
            send_mail(subject, message, cls.FROM_EMAIL, [admin_email], fail_silently=True)
        except Exception as e:
            print(f"Failed to send low stock alert email: {e}")
            
    @classmethod
    def send_lab_result_approved(cls, email, sample_id, product):
        subject = "Lab Result Approved"
        message = f"Hello,\n\nThe lab result for Sample '{sample_id}' (Product: {product}) has been approved.\n\nRegards\nGSFC LIMS System"
        try:
            send_mail(subject, message, cls.FROM_EMAIL, [email], fail_silently=True)
        except Exception as e:
            print(f"Failed to send lab approval email: {e}")

    @classmethod
    def send_high_severity_anomaly_alert(cls, admin_emails, sample_data):
        subject = f"LAB QUALITY ALERT — Sample {sample_data.get('sample_id', 'Unknown')}"
        message = (
            f"HIGH SEVERITY ANOMALY DETECTED\n\n"
            f"Sample ID: {sample_data.get('sample_id')}\n"
            f"Product: {sample_data.get('product')}\n"
            f"Moisture: {sample_data.get('moisture')}%\n"
            f"Purity: {sample_data.get('purity')}%\n"
            f"Analyst: {sample_data.get('analyst')}\n\n"
            f"Reason: {sample_data.get('reason')}\n\n"
            f"Please review this immediately in the LIMS Dashboard Quality Insights panel.\n\n"
            f"Regards,\nGSFC LIMS System"
        )
        try:
            send_mail(subject, message, cls.FROM_EMAIL, admin_emails, fail_silently=True)
        except Exception as e:
            print(f"Failed to send High Severity Anomaly Email: {e}")

    @classmethod
    def send_anomaly_review_notification(cls, admin_emails, sample_id, action_taken, analyst, reason):
        subject = f"Anomaly Reviewed: {action_taken} - Sample {sample_id}"
        message = (
            f"Hello,\n\n"
            f"The anomaly for Sample ID {sample_id} was marked as {action_taken} by {analyst}.\n\n"
            f"Original Anomaly Reason: {reason}\n\n"
            f"Regards,\nGSFC LIMS System"
        )
        try:
            send_mail(subject, message, cls.FROM_EMAIL, admin_emails, fail_silently=True)
        except Exception as e:
            print(f"Failed to send Anomaly Review Email: {e}")


# ==========================================
# 2. ALERT SERVICE
# ==========================================

class AlertService:
    @classmethod
    def get_active_alerts(cls):
        alerts = []
        INVENTORY_MASTER_PATH = os.path.join(os.getcwd(), "inventory_master.xlsx")
        
        if not os.path.exists(INVENTORY_MASTER_PATH):
            return alerts
            
        try:
            wb = load_workbook(INVENTORY_MASTER_PATH, data_only=True)
            ws = wb.active
            
            # Assuming columns:
            # 0: material_id, 1: material_name, 2: unit, 3: current_balance, 4: min_threshold
            for row in ws.iter_rows(min_row=2, values_only=True):
                material = row[1]
                if not material:
                    continue
                    
                current_balance = float(row[3]) if row[3] is not None else 0
                threshold = float(row[4]) if row[4] is not None else 0
                
                if current_balance < threshold:
                    alerts.append({
                        "id": f"ALT-INV-{row[0]}",
                        "material": material,
                        "current_balance": current_balance,
                        "threshold": threshold,
                        "type": "LOW_STOCK",
                        "message": f"{material} stock is {current_balance} (Below {threshold})"
                    })
                    
        except Exception as e:
            print(f"Failed to generate alerts: {e}")
            
        return alerts
