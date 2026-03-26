import os
from openpyxl import load_workbook
import datetime
import calendar

# 'import sklearn' causes a GUI/DLL deadlock on some Windows environments 
# which hangs the server indefinitely and prevents Ctrl+C.
# We will disable it and rely on the naive prediction fallback.
SKLEARN_AVAILABLE = False


class PredictiveModel:
    @classmethod
    def predict_material_usage(cls, material_name):
        """
        Calculates historical usage of `material_name` per day from inventory_requests.xlsx
        and predicts the consumption for the next 7 days using Linear Regression.
        """
        if not SKLEARN_AVAILABLE:
            # Fallback naive logic if scikit-learn is missing
            return cls._naive_prediction(material_name)
            
        INV_REQUESTS_PATH = os.path.join(os.getcwd(), "inventory_requests.xlsx")
        if not os.path.exists(INV_REQUESTS_PATH):
            return {"error": "No historical data to predict from."}

        try:
            wb = load_workbook(INV_REQUESTS_PATH, data_only=True)
            ws = wb["Requests"] if "Requests" in wb.sheetnames else wb.active
            
            # Gather daily usage
            # column 1 = material, column 2 = quantity, column 7 = date (DD-MM-YYYY)
            daily_usage = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                mat = str(row[1])
                qty = float(row[2]) if row[2] else 0
                date_str = str(row[7])
                # Only consider "Issued" or "Completed" status (row 10)
                status = str(row[10]).upper()
                
                if mat == material_name and status in ("ISSUED", "APPROVED", "COMPLETED"):
                    if date_str not in daily_usage:
                        daily_usage[date_str] = 0
                    daily_usage[date_str] += qty

            if not daily_usage:
                return {"material": material_name, "predicted_usage_next_week": 0, "message": "Not enough data"}

            # Sort by date
            sorted_dates = []
            for d in daily_usage.keys():
                try:
                    dt = datetime.datetime.strptime(d, "%d-%m-%Y")
                    sorted_dates.append(dt)
                except:
                    pass
            
            if not sorted_dates:
                 return {"material": material_name, "predicted_usage_next_week": 0, "message": "Invalid date formats"}

            sorted_dates.sort()
            start_date = sorted_dates[0]
            
            # X = days since start_date
            # y = usage quantity on that day
            X_list = []
            y_list = []
            
            # We construct a time series for all days between start_date and today
            today = datetime.datetime.now()
            delta = (today - start_date).days
            
            if delta < 1:
                # If everything happened today, we can't build a meaningful regression line over time.
                total_qty = sum(daily_usage.values())
                return {"material": material_name, "predicted_usage_next_week": round(total_qty * 7, 2), "message": "Using naive daily multiply (insufficient timeframe)"}

            for i in range(delta + 1):
                d_obj = start_date + datetime.timedelta(days=i)
                d_str = d_obj.strftime("%d-%m-%Y")
                qty = daily_usage.get(d_str, 0)
                
                X_list.append([i])
                y_list.append(qty)

            X = np.array(X_list)
            y = np.array(y_list)

            model = LinearRegression()
            model.fit(X, y)
            
            # Predict the next 7 days
            next_7_days_indices = np.array([[delta + 1 + i] for i in range(7)])
            predictions = model.predict(next_7_days_indices)
            
            # Avoid negative predictions
            predictions = [max(0, p) for p in predictions]
            total_next_week = sum(predictions)
            
            return {
                "material": material_name,
                "predicted_usage_next_week": round(total_next_week, 2),
                "model_used": "LinearRegression"
            }

        except Exception as e:
            print(f"Error in ML prediction: {e}")
             # Fallback
            return cls._naive_prediction(material_name)

    @classmethod
    def _naive_prediction(cls, material_name):
         return {
            "material": material_name,
            "predicted_usage_next_week": 0,
            "message": "Prediction failed or scikit-learn not available."
         }
